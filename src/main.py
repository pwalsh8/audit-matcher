import streamlit as st
import pandas as pd
from matcher import (match_documents, extract_text_from_pdf, extract_primary_amount, 
                    process_files, extract_amounts_from_text)
from utils import (ensure_output_directory, save_matches_to_excel, cleanup_output_images, load_and_validate_excel, 
                  PDFHandler, CommonUtils)
import logging
import tempfile
import time
from typing import List, Dict, Any
from pathlib import Path
from PIL import Image
import io  # Add missing import
from pdf2image import convert_from_path
from openpyxl import Workbook
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from constants import SUPPORT_CATEGORIES, OUTPUT_FOLDER, PREVIEW_FILE, SUMMARY_FILE
import shutil  # Add this import

# Configure logging with a more standardized setup
logging.basicConfig(
    filename='app.log',
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

# Remove duplicate function and use CommonUtils directly
format_currency = CommonUtils.format_currency
get_file_path = CommonUtils.get_file_path
normalize_amount = CommonUtils.normalize_amount

def handle_error(error_message):
    """
    Handle errors and logging.
    
    Args:
        error_message: The error message to be logged and displayed
    """
    logger.error(error_message)
    st.error(error_message)
    st.write("Please check your files and try again")

def show_match_confirmation(matches):
    st.header("Confirm Matches")
    st.write("Please review and confirm the matches for each selection.")
    
    # Use st.expander for each match to make it cleaner
    for match in matches:
        with st.expander(f"Selection {match['Selection ID']}", expanded=True):
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("Selection Details")
                st.write(f"Amount: ${float(match['Selection Amount']):,.2f}")
                st.write(f"ID: {match['Selection ID']}")
                
            with col2:
                st.subheader("Matched PDF")
                st.write(f"File: {match['PDF Name']}")
                st.write(f"Amount: ${float(match['PDF Amount']):,.2f}")
                st.write(f"Match Type: {match['Match Type']}")
                
            # Add confirm/reject buttons
            col3, col4 = st.columns(2)
            with col3:
                st.button("✓ Confirm Match", key=f"confirm_{match['Selection ID']}")
            with col4:
                st.button("❌ Reject Match", key=f"reject_{match['Selection ID']}")
                
    # Add a final confirmation button
    if st.button("✅ Confirm All Matches", type="primary"):
        st.success("All matches confirmed!")
        return matches
        
    return None

def save_uploaded_pdf(pdf_file) -> Path:
    """Save uploaded PDF to temporary file for processing"""
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
            tmp.write(pdf_file.getvalue())
            return Path(tmp.name)
    except Exception as e:
        logger.error(f"Error saving PDF file: {str(e)}")
        raise

def preview_pdf_amounts(pdf_files: List) -> pd.DataFrame:
    """Extract and preview amounts from uploaded PDFs"""
    preview_data = []
    
    for pdf_file in pdf_files:
        try:
            # Safe filename extraction
            if isinstance(pdf_file, (str, Path)):
                filename = str(pdf_file)
            elif hasattr(pdf_file, 'name'):
                filename = pdf_file.name
            else:
                filename = str(pdf_file)
                
            temp_path = get_file_path(pdf_file)
            text = extract_text_from_pdf(temp_path)
            amount = extract_primary_amount(text)
            
            if amount:
                raw_amount = str(amount)
                normalized = normalize_amount(amount)
                
                preview_data.append({
                    "PDF Name": filename,
                    "Raw Amount": f"${raw_amount}",
                    "Normalized Amount": f"${normalized:,.2f}",
                    "Status": "✓"
                })
            else:
                preview_data.append({
                    "PDF Name": filename,
                    "Raw Amount": "No amount found",
                    "Normalized Amount": "N/A",
                    "Status": "⚠️"
                })
            
            if isinstance(pdf_file, st.runtime.uploaded_file_manager.UploadedFile):
                temp_path.unlink()
                
        except Exception as e:
            logger.error(f"Error processing PDF: {str(e)}")
            preview_data.append({
                "PDF Name": filename if 'filename' in locals() else "Unknown",
                "Raw Amount": "Error",
                "Normalized Amount": "Error",
                "Status": "❌"
            })
    
    return pd.DataFrame(preview_data)

def preview_match(selection: Dict, pdf_info: Dict):
    """Show side-by-side preview of selection and PDF for manual confirmation"""
    match_container = st.container()
    with match_container:
        col1, col2 = st.columns(2)
        
        with col1:
            st.write("### Selection Details")
            st.write(f"ID: {selection['Selection ID']}")
            st.write(f"Amount: {selection['Selection Amount']}")
            for key, value in selection['Selection Data'].items():
                if key not in ['Selection ID', 'amount_column']:
                    st.write(f"{key}: {value}")

        with col2:
            st.write("### PDF Preview")
            st.write(f"Filename: {pdf_info['PDF Name']}")
            st.write(f"Found Amount: {pdf_info['PDF Amount']}")
            
            # Show PDF preview with error handling
            if pdf_info.get('Matched Pages'):
                for i, page_path in enumerate(pdf_info['Matched Pages']):
                    try:
                        img_path = Path(page_path)
                        if img_path.exists():
                            with Image.open(img_path) as img:
                                st.image(img, caption=f"Page {i+1}", use_column_width=True)
                        else:
                            st.warning(f"Image not found: {page_path}")
                    except Exception as e:
                        st.error(f"Error loading image: {str(e)}")
                        logger.error(f"Image load error: {str(e)}")

def interactive_matching(df: pd.DataFrame, pdf_files: List, unique_id_column: str, amount_column: str):
    """Interactive matching process with preview and confirmation"""
    
    # Initialize session state for matches if not exists
    if 'confirmed_matches' not in st.session_state:
        st.session_state.confirmed_matches = []
    
    with st.spinner("Processing matches..."):
        # Get potential matches
        result = process_files(df, unique_id_column, amount_column, pdf_files)
        
        if "error" in result:
            st.error(f"An error occurred: {result['error']}")
            return
            
        matches = result.get("matches", [])
        
        if not matches:
            st.warning("No potential matches found.")
            return
            
        # Create tabs for different views
        tab1, tab2 = st.tabs(["Review Matches", "Summary"])
        
        with tab1:
            st.header("Review Matches")
            st.write("Please review and confirm each match below.")
            
            for i, match in enumerate(matches):
                with st.expander(f"Selection {match['Selection ID']}", expanded=True):
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.subheader("Selection Details")
                        st.metric("Amount", match['Selection Amount'])
                        st.write(f"ID: {match['Selection ID']}")
                    
                    with col2:
                        st.subheader("PDF Match")
                        st.write(f"File: {match['PDF Name']}")
                        st.metric("Amount", match['PDF Amount'])
                        st.write(f"Match Type: {match['Match Type']}")
                    
                    # Add action buttons
                    col3, col4, col5 = st.columns(3)
                    with col3:
                        if st.button("✓ Confirm", key=f"confirm_{i}_{match['Selection ID']}"):
                            if match not in st.session_state.confirmed_matches:
                                st.session_state.confirmed_matches.append(match)
                                st.success("Match confirmed!")
                    
                    with col4:
                        if st.button("❌ Reject", key=f"reject_{i}_{match['Selection ID']}"):
                            if match in st.session_state.confirmed_matches:
                                st.session_state.confirmed_matches.remove(match)
                            st.info("Match rejected")
                            
                    with col5:
                        if st.button("👁️ View PDF", key=f"view_{i}_{match['Selection ID']}"):
                            # Show PDF preview if available
                            if match.get('Matched Pages'):
                                for i, page in enumerate(match['Matched Pages']):
                                    st.image(page, caption=f"Page {i+1}")
        
        with tab2:
            st.header("Matching Summary")
            st.metric("Total Matches", len(matches))
            st.metric("Confirmed Matches", len(st.session_state.confirmed_matches))
            
            if st.session_state.confirmed_matches:
                if st.button("📥 Download Results", type="primary"):
                    try:
                        # Ensure output directory exists
                        output_dir = ensure_output_directory()
                        output_path = output_dir / "matching_results.xlsx"
                        
                        # Save Excel file with user labels
                        user_labels = {
                            "Unique ID Column": unique_id_column,
                            "Amount Column": amount_column,
                            "Processing Date": time.strftime("%Y-%m-%d %H:%M:%S"),
                            "Total Matches": len(st.session_state.confirmed_matches)
                        }
                        
                        save_matches_to_excel(st.session_state.confirmed_matches, output_path, user_labels)
                        
                        # Create download button
                        with open(output_path, "rb") as file:
                            excel_data = file.read()
                        
                        st.download_button(
                            label="📥 Download Matching Results",
                            data=excel_data,
                            file_name="matching_results.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_results"
                        )
                        
                        st.success("✓ Excel report generated successfully!")
                        
                    except Exception as e:
                        logger.error(f"Error preparing Excel download: {e}")
                        st.error("Error generating Excel file. Please try again.")

def _advance_match(state: Dict):
    """Helper to advance to next match"""
    state['current_index'] += 1
    if state['current_index'] >= state['total_matches']:
        state['completed'] = True
    st.rerun()  # Use st.rerun() instead of st.experimental_rerun()

def convert_pdf_to_images(pdf_path: Path) -> List[Path]:
    """Convert PDF pages to images and return list of image paths"""
    try:
        images = convert_from_path(pdf_path)
        image_paths = []
        for i, image in enumerate(images):
            image_path = pdf_path.with_name(f"{pdf_path.stem}_page_{i+1}.png")
            image.save(image_path, "PNG")
            image_paths.append(image_path)
        return image_paths
    except Exception as e:
        logger.error(f"Error converting PDF to images: {str(e)}")
        raise

def embed_images_in_excel(sheet, image_paths: List[Path], start_row: int):
    """Embed images into an Excel sheet starting from a specific row"""
    for i, image_path in enumerate(image_paths):
        img = ExcelImage(str(image_path))
        sheet.add_image(img, f'A{start_row + i}')
        # Optionally, remove the image file after embedding
        image_path.unlink(missing_ok=True)

def save_matches_to_excel_with_images(matches: List[Dict], output_path: Path, user_labels: Dict):
    """Save matches to an Excel file with embedded images"""
    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    
    # Add report header
    summary_sheet['A1'] = "Audit Matcher Results"
    summary_sheet['A1'].font = openpyxl.styles.Font(bold=True, size=14)
    summary_sheet['A2'] = f"Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Add statistics
    row = 4
    summary_sheet[f'A{row}'] = "Statistics"
    summary_sheet[f'A{row}'].font = openpyxl.styles.Font(bold=True)
    row += 1
    
    for label, value in user_labels.items():
        summary_sheet[f'A{row}'] = label
        summary_sheet[f'B{row}'] = value
        row += 1
    
    for match in matches:
        sheet_name = str(match['Selection ID'])[:31]  # Excel sheet name limit
        sheet = wb.create_sheet(title=sheet_name)
        
        # Write selection data
        for col, (key, value) in enumerate(match['Selection Data'].items(), start=1):
            sheet.cell(row=1, column=col, value=f"{key}: {value}")
        
        # Add blank row for separation
        sheet.cell(row=2, column=1, value="")
        
        # Convert PDF to images and embed in Excel
        pdf_path = Path(match['PDF Path'])
        image_paths = convert_pdf_to_images(pdf_path)
        embed_images_in_excel(sheet, image_paths, start_row=3)
    
    # Remove default sheet created by openpyxl
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    wb.save(output_path)
    logger.info(f"Excel file saved successfully to {output_path}")

def show_matching_summary(confirmed_matches: List[Dict], unique_id_column: str, amount_column: str) -> None:
    """Display summary of matching results"""
    st.write("### Matching Summary")
    
    # Show summary stats in columns
    col1, col2 = st.columns(2)
    with col1:
        st.write(f"Total Confirmed: {len(confirmed_matches)}")
        if confirmed_matches:
            st.write("### Confirmed Matches")
            for match in confirmed_matches:
                with st.expander(f"Selection {match['Selection ID']} - {match['Selection Amount']}"):
                    st.write(f"Matched PDF: {match['PDF Name']}")
                    st.write(f"PDF Amount: {match['PDF Amount']}")
                    if match.get('Matched Pages'):
                        try:
                            for i, page_path in enumerate(match['Matched Pages']):
                                img_path = Path(page_path)
                                if img_path.exists():
                                    with Image.open(img_path) as img:
                                        st.image(img, caption=f"Page {i+1}", use_column_width=True)
                        except Exception as e:
                            st.error(f"Error loading PDF preview: {str(e)}")
    
    with col2:
        # Get skipped matches from the confirmed matches that have skipped=True
        skipped_matches = [m for m in confirmed_matches if m.get('skipped', False)]
        st.write(f"Total Skipped: {len(skipped_matches)}")
        if skipped_matches:
            for skip in skipped_matches:
                st.write(f"- Selection {skip['Selection ID']} ({skip['Selection Amount']})")

def show_matching_summary(confirmed_matches: List[Dict], unique_id_column: str, amount_column: str) -> None:
    """Display summary of matching results"""
    st.write("### Matching Summary")
    
    # Show summary stats in columns
    col1, col2 = st.columns(2)
    with col1:
        st.write(f"Total Confirmed: {len(confirmed_matches)}")
        if confirmed_matches:
            st.write("### Confirmed Matches")
            for match in confirmed_matches:
                with st.expander(f"Selection {match['Selection ID']} - {match['Selection Amount']}"):
                    st.write(f"Matched PDF: {match['PDF Name']}")
                    st.write(f"PDF Amount: {match['PDF Amount']}")
                    if match.get('Matched Pages'):
                        try:
                            for i, page_path in enumerate(match['Matched Pages']):
                                img_path = Path(page_path)
                                if img_path.exists():
                                    with Image.open(img_path) as img:
                                        st.image(img, caption=f"Page {i+1}", use_column_width=True)
                        except Exception as e:
                            st.error(f"Error loading PDF preview: {str(e)}")
    
    with col2:
        # Get skipped matches from the confirmed matches that have skipped=True
        skipped_matches = [m for m in confirmed_matches if m.get('skipped', False)]
        st.write(f"Total Skipped: {len(skipped_matches)}")
        if skipped_matches:
            for skip in skipped_matches:
                st.write(f"- Selection {skip['Selection ID']} ({skip['Selection Amount']})")
    
    # Export section
    if confirmed_matches:
        st.write("---")
        st.write("### Export Results")
        
        try:
            # Ensure output directory exists
            output_dir = ensure_output_directory()
            output_path = output_dir / "matching_results.xlsx"
            
            # Save Excel file with user labels
            user_labels = {
                "Unique ID Column": unique_id_column,
                "Amount Column": amount_column,
                "Processing Date": time.strftime("%Y-%m-%d %H:%M:%S"),
                "Total Matches": len(confirmed_matches)
            }
            
            save_matches_to_excel(confirmed_matches, output_path, user_labels)
            
            # Create download button
            with open(output_path, "rb") as file:
                excel_data = file.read()
            
            st.download_button(
                label="📥 Download Results",
                data=excel_data,
                file_name="matching_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            st.success("✓ Excel report generated successfully!")
            
        except Exception as e:
            logger.error(f"Error preparing Excel download: {e}")
            st.error("Error generating Excel file. Please try again.")

def process_matched_files(matches: List[Dict]) -> None:
    """Process matched files and create organized output"""
    try:
        # Create base output directory
        output_dir = Path(OUTPUT_FOLDER)
        output_dir.mkdir(exist_ok=True)
        
        # Create folders for each selection
        selection_folders = create_selection_folders([m['Selection ID'] for m in matches])
        
        # Move PDFs to appropriate folders
        for match in matches:
            selection_id = match['Selection ID']
            pdf_file = match['PDF Name']
            target_folder = selection_folders[selection_id]
            
            # Copy PDF to selection folder
            shutil.copy2(pdf_file, target_folder / pdf_file.name)
        
        # Create summary Excel file
        create_support_summary(matches, output_dir / SUMMARY_FILE)
        
        # Create preview table
        preview_df = create_preview_table(matches, output_dir / PREVIEW_FILE)
        
        return {
            'status': 'success',
            'preview': preview_df,
            'output_dir': output_dir
        }
        
    except Exception as e:
        logger.error(f"Error processing matched files: {e}")
        return {'status': 'error', 'error': str(e)}

def show_preview_interface(preview_df: pd.DataFrame) -> None:
    """Show preview interface with support categories"""
    st.write("### Support Preview")
    st.dataframe(preview_df)
    
    # Download buttons for Excel files
    col1, col2 = st.columns(2)
    with col1:
        with open(Path(OUTPUT_FOLDER) / PREVIEW_FILE, 'rb') as f:
            st.download_button(
                "Download Preview Table",
                f,
                file_name=PREVIEW_FILE,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
    with col2:
        with open(Path(OUTPUT_FOLDER) / SUMMARY_FILE, 'rb') as f:
            st.download_button(
                "Download Full Summary",
                f,
                file_name=SUMMARY_FILE,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

def create_selection_folders(selection_ids: List[str]) -> Dict[str, Path]:
    """Create folders for each selection ID and return mapping of ID to folder path"""
    selection_folders = {}
    for selection_id in selection_ids:
        folder_path = Path(OUTPUT_FOLDER) / str(selection_id)
        folder_path.mkdir(parents=True, exist_ok=True)
        selection_folders[selection_id] = folder_path
    return selection_folders

def create_support_summary(matches: List[Dict], output_path: Path) -> None:
    """Create Excel summary of matched documents"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Support Summary"

    # Define headers
    headers = ['Selection ID', 'Selection Amount', 'PDF Name', 'PDF Amount', 
              'PDF Type', 'Match Type', 'Status']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Add match data
    for row, match in enumerate(matches, 2):
        ws.cell(row=row, column=1, value=match['Selection ID'])
        ws.cell(row=row, column=2, value=match['Selection Amount'])
        ws.cell(row=row, column=3, value=match['PDF Name'])
        ws.cell(row=row, column=4, value=match['PDF Amount'])
        ws.cell(row=row, column=5, value=match.get('PDF Type', 'Unknown'))
        ws.cell(row=row, column=6, value=match['Match Type'])
        ws.cell(row=row, column=7, value='Matched')

    wb.save(output_path)

def create_preview_table(matches: List[Dict], output_path: Path) -> pd.DataFrame:
    """Create preview table of matches and save to Excel"""
    preview_data = []
    
    for match in matches:
        preview_data.append({
            'Selection ID': match['Selection ID'],
            'Selection Amount': match['Selection Amount'],
            'PDF Name': match['PDF Name'],
            'Support Type': match.get('PDF Type', 'Unknown'),
            'Status': 'Matched'
        })
    
    df = pd.DataFrame(preview_data)
    df.to_excel(output_path, index=False)
    return df

def main():
    st.title("Audit Matcher")
    st.write("Welcome to Audit Matcher!")

    # Initialize session state if needed
    if 'app_state' not in st.session_state:
        st.session_state.app_state = {
            'matching_started': False,
            'df': None,
            'pdf_files': None,
            'columns': {'id': None, 'amount': None}
        }

    # File uploaders section
    st.write("### Step 1: Upload Excel File")
    selections_file = st.file_uploader(
        "Upload Selections Excel",
        type=['xlsx'],
        help="Select your audit selections Excel file"
    )

    if selections_file:
        try:
            # Load Excel file
            df = pd.read_excel(selections_file)
            if df.empty:
                st.error("The uploaded Excel file is empty")
                return

            st.success("✓ Excel file loaded successfully")
            st.session_state.app_state['df'] = df
            
            # Display raw data
            st.write("### Raw Data Preview:")
            st.dataframe(df.head())
            
            # Column selection
            columns = df.columns.tolist()
            unique_id_column = st.selectbox("Select ID Column", options=columns)
            amount_column = st.selectbox("Select Amount Column", options=columns)
            
            if amount_column:
                # Simple amount preview
                preview_df = pd.DataFrame({
                    'ID': df[unique_id_column],
                    'Amount': df[amount_column]
                })
                st.dataframe(preview_df)

            # PDF upload section
            st.write("### Step 2: Upload PDF Files")
            pdf_files = st.file_uploader(
                "Upload PDF Files",
                type=['pdf'],
                accept_multiple_files=True,
                help="Select one or more PDF files to match against the Excel data"
            )

            if pdf_files:
                st.session_state.app_state['pdf_files'] = pdf_files
                st.session_state.app_state['columns'] = {
                    'id': unique_id_column,
                    'amount': amount_column
                }
                
                st.success(f"✓ {len(pdf_files)} PDF file(s) uploaded successfully")
                
                # Show PDF previews
                with st.spinner("Extracting amounts from PDFs..."):
                    preview_df = preview_pdf_amounts(pdf_files)
                    st.write("### PDF Amount Preview")
                    st.dataframe(preview_df)

                # Add support category selection
                category = st.selectbox(
                    "Select Support Type",
                    options=SUPPORT_CATEGORIES
                )

                # Start matching button
                if st.button("Start Matching") or st.session_state.app_state.get('matching_started'):
                    st.session_state.app_state['matching_started'] = True
                    matches = match_documents(df, unique_id_column, amount_column, pdf_files)

                    # Add support category to matches
                    for match in matches:
                        match['PDF Type'] = category

                    # Process matched files
                    result = process_matched_files(matches)

                    if result['status'] == 'success':
                        show_preview_interface(result['preview'])
                        st.success(f"Files organized in: {result['output_dir']}")
                    else:
                        st.error(f"Error: {result.get('error')}")

                # Add interactive matching button
                if st.button("Start Interactive Matching"):
                    with st.spinner("Matching documents..."):
                        matches = interactive_matching(df, pdf_files, unique_id_column, amount_column)
                        
                        if matches:
                            st.success("✓ Matching completed!")
                            
                            # Show results in tabs
                            tab1, tab2 = st.tabs(["Match Details", "Summary"])
                            
                            with tab1:
                                for match in matches:
                                    with st.expander(f"Selection {match['Selection ID']} - {match['PDF Name']}"):
                                        col1, col2 = st.columns(2)
                                        with col1:
                                            st.write("### Selection Details")
                                            st.write(f"Amount: {match['Selection Amount']}")
                                        with col2:
                                            st.write("### PDF Details")
                                            st.write(f"Amount: {match['PDF Amount']}")
                                            st.write(f"Match Type: {match['Match Type']}")
                                            
                            with tab2:
                                st.write(f"Total Matches: {len(matches)}")
                                # Add download button for results
                                output_path = "matching_results.xlsx"
                                save_matches_to_excel(matches, output_path)
                                
                                with open(output_path, "rb") as file:
                                    st.download_button(
                                        label="📥 Download Results",
                                        data=file,
                                        file_name=output_path,
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                    )

        except Exception as e:
            handle_error(f"Error: {str(e)}")
            logger.exception("Error in main:")
    else:
        st.info("Please upload an Excel file to begin")

    # Debug section
    if st.checkbox("Show Debug Info"):
        st.write("Session State:", st.session_state.app_state)

if __name__ == "__main__":
    main()