"""
Utility functions for audit matching:
1. Load and validate Excel files
2. Process PDF files safely
3. Format currency amounts
4. Handle errors and logging
"""

import pandas as pd
import pdfplumber  # Changed from PyPDF2 to pdfplumber to match matcher.py
import logging
from decimal import Decimal, InvalidOperation
from typing import Optional, List, Dict, Union, Any
from pathlib import Path
import streamlit as st
import openpyxl
from openpyxl.drawing.image import Image as OpenPyxlImage
from pdf2image import convert_from_path
import tempfile
import os
import shutil
import time
from constants import SUPPORT_CATEGORIES, OUTPUT_FOLDER, EXCEL_HEADERS

# Configure logging
logging.basicConfig(
    filename='app.log',
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

def ensure_output_directory() -> Path:
    """Ensure output directory exists and return absolute Path"""
    # Get absolute base path - this should be project root
    base_dir = Path.cwd()
    if 'src' in base_dir.parts:
        base_dir = base_dir.parent
    
    output_dir = base_dir / "output_images"
    temp_dir = output_dir / "temp"
    
    # Create both directories if they don't exist
    output_dir.mkdir(parents=True, exist_ok=True)
    temp_dir.mkdir(parents=True, exist_ok=True)
    
    # Return resolved absolute path
    return output_dir.resolve()

def load_and_validate_excel(file_path: Union[str, Path, st.runtime.uploaded_file_manager.UploadedFile]) -> Optional[pd.DataFrame]:
    """Load and validate an Excel file containing audit selections."""
    try:
        df = pd.read_excel(file_path)
        
        # Remove any unnamed columns
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        
        # Create Selection ID if not present
        if 'Selection ID' not in df.columns:
            df['Selection ID'] = range(1, len(df) + 1)
        
        # Clean and validate numeric columns
        for col in df.columns:
            if df[col].dtype == object:  # Only process string columns
                try:
                    # Remove currency symbols and commas
                    cleaned = df[col].astype(str).replace('[\$,]', '', regex=True)
                    # Try converting to float
                    if cleaned.str.match(r'^-?\d*\.?\d*$').all():
                        df[col] = pd.to_numeric(cleaned, errors='coerce')
                except Exception as e:
                    logger.debug(f"Column {col} not numeric: {e}")
                    continue
            
        logger.info(f"Loaded Excel with {len(df)} rows")
        return df
        
    except Exception as e:
        log_error(f"Error loading Excel file: {str(e)}")
        return None

def log_error(message: str, exception: Optional[Exception] = None) -> None:
    """
    Log error message and optionally the exception.
    
    Args:
        message: Error message to log
        exception: Optional exception object
    """
    logger.error(message)
    if exception:
        logger.error(f"Exception details: {str(exception)}", exc_info=True)
    
def validate_dataframe(df: pd.DataFrame, required_columns: List[str]) -> bool:
    """
    Validate DataFrame has required columns and data.
    
    Args:
        df: DataFrame to validate
        required_columns: List of required column names
    
    Returns:
        bool: True if valid, False otherwise
    """
    try:
        if df.empty:
            logger.error("DataFrame is empty")
            return False
            
        missing_columns = set(required_columns) - set(df.columns)
        if missing_columns:
            logger.error(f"Missing columns: {missing_columns}")
            return False
            
        return True
    except Exception as e:
        log_error("Error validating DataFrame", e)
        return False

def save_upload_file(uploadedfile: st.runtime.uploaded_file_manager.UploadedFile) -> Optional[Path]:
    """
    Save uploaded file to temporary location.
    
    Args:
        uploadedfile: Streamlit uploaded file
    
    Returns:
        Optional[Path]: Path to saved file or None if error occurs
    """
    try:
        temp_dir = Path("temp")
        temp_dir.mkdir(exist_ok=True)
        
        temp_path = temp_dir / uploadedfile.name
        with open(temp_path, "wb") as f:
            f.write(uploadedfile.getbuffer())
            
        return temp_path
    except Exception as e:
        log_error(f"Error saving uploaded file: {str(e)}")
        return None

def cleanup_temp_files(temp_dir: Union[str, Path] = "temp") -> None:
    """
    Clean up temporary files.
    
    Args:
        temp_dir: Directory containing temporary files
    """
    try:
        temp_path = Path(temp_dir)
        if temp_path.exists():
            shutil.rmtree(temp_path)
    except Exception as e:
        log_error(f"Error cleaning up temp files: {str(e)}")

def save_matches_to_excel(matches: List[Dict[str, Any]], output_path: Union[str, Path], user_labels: Dict[str, str]) -> None:
    """Save matching results to an Excel file with embedded PDF images."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = None
    try:
        wb = openpyxl.Workbook()
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        
        # Set column widths
        summary_sheet.column_dimensions['A'].width = 30
        summary_sheet.column_dimensions['B'].width = 30
        
        # Add report header
        summary_sheet['A1'] = "Audit Matcher Results"
        summary_sheet['A1'].font = openpyxl.styles.Font(bold=True, size=14)
        summary_sheet['A2'] = f"Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Add statistics
        row = 4
        summary_sheet[f'A{row}'] = "Statistics"
        summary_sheet[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1
        
        for label, value in user_labels.items():
            summary_sheet[f'A{row}'] = label
            summary_sheet[f'B{row}'] = value
            row += 1
        
        # Process each match on its own tab
        for match in matches:
            try:
                # Create sheet with selection ID as name
                sheet_name = str(match.get('Selection ID', ''))[:31]  # Excel limit
                ws = wb.create_sheet(title=sheet_name)
                logger.debug(f"Processing match for Selection ID: {sheet_name}")
                
                # Set column widths
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 40
                
                # Add selection data header
                cell = ws.cell(row=1, column=1, value="Selection Details")
                cell.font = openpyxl.styles.Font(bold=True, size=12)
                
                # Add selection data
                current_row = 2
                selection_data = match.get('Selection Data', {})
                for key, value in selection_data.items():
                    ws.cell(row=current_row, column=1, value=key)
                    ws.cell(row=current_row, column=2, value=str(value))  # Fix: Correct parentheses
                    current_row += 1
                
                # Add spacing
                current_row += 1
                
                # Add match details
                cell = ws.cell(row=current_row, column=1, value="Match Details")
                cell.font = openpyxl.styles.Font(bold=True)
                current_row += 1
                
                # Add match information
                match_info = [
                    ("PDF Filename", match.get('PDF Name', '')),
                    ("Selection Amount", match.get('Selection Amount', '')),
                    ("PDF Amount", match.get('PDF Amount', '')),
                    ("Match Type", match.get('Match Type', ''))
                ]
                
                for label, value in match_info:
                    ws.cell(row=current_row, column=1, value=label)
                    ws.cell(row=current_row, column=2, value=value)
                    current_row += 1
                
                current_row += 1
                
                # Add PDF images if available
                matched_pages = match.get('Matched Pages', [])
                if matched_pages:
                    cell = ws.cell(row=current_row, column=1, value="PDF Pages")
                    cell.font = openpyxl.styles.Font(bold=True)
                    current_row += 1
                    
                    for i, page_path in enumerate(matched_pages, 1):
                        try:
                            # Resize and add image
                            if Path(page_path).exists():
                                img = OpenPyxlImage(page_path)
                                
                                # Set image size in cells
                                ws.row_dimensions[current_row].height = 300
                                img.width = 600
                                img.height = 400
                                
                                # Position image
                                img.anchor = f'B{current_row}'
                                ws.add_image(img)
                                
                                ws.cell(row=current_row, column=1, value=f"Page {i}")
                                current_row += 22  # Space for image
                                
                        except Exception as e:
                            logger.error(f"Error adding image {page_path}: {e}")
                            ws.cell(row=current_row, column=1, value=f"Error adding page {i}")
                            current_row += 1
                else:
                    ws.cell(row=current_row, column=1, value="No PDF pages found")
            
            except Exception as e:
                logger.error(f"Error processing match {match.get('Selection ID')}: {e}")
                continue

        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel file saved successfully to {output_path}")
        
    except Exception as e:
        logger.error(f"Error creating Excel file: {e}")
        raise
    finally:
        if wb:
            try:
                wb.close()
            except Exception as e:
                logger.error(f"Error closing workbook: {e}")

def cleanup_output_images() -> None:
    output_dir = "output_images"
    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)  # Caution: This deletes the entire directory and its contents

def resize_pdf_image(image_path: Union[str, Path], max_size: tuple = (800, 600)) -> str:
    """
    Resize a PDF page image while maintaining aspect ratio.
    """
    from PIL import Image
    
    image_path = Path(image_path).resolve()  # Get absolute path
    if not image_path.exists():
        logger.error(f"Source image not found: {image_path}")
        raise FileNotFoundError(f"Source image not found: {image_path}")

    try:
        # Ensure output directory exists
        output_dir = ensure_output_directory()
        
        with Image.open(image_path) as img:
            # Calculate new dimensions maintaining aspect ratio
            ratio = min(max_size[0]/img.width, max_size[1]/img.height)
            new_size = (int(img.width * ratio), int(img.height * ratio))
            
            # Resize image
            resized = img.resize(new_size, Image.Resampling.LANCZOS)
            
            # Generate safe output filename
            safe_name = Path(image_path).stem.replace('.', '_').replace(' ', '_')
            output_path = output_dir / f"{safe_name}_resized.png"
            
            # Save with absolute path
            resized.save(str(output_path.resolve()), "PNG", optimize=True)
            
            if not output_path.exists():
                raise IOError(f"Failed to create resized image at {output_path}")
            
            logger.info(f"Successfully saved resized image to: {output_path}")
            return str(output_path.resolve())
            
    except Exception as e:
        logger.error(f"Error resizing image {image_path}: {e}")
        raise

def generate_output(matches: List[Dict[str, Any]], output_file: str='matching_results.xlsx'):
    """
    Create DataFrame with columns:
    - Selection ID
    - Excel Amount
    - PDF Filename
    - PDF Amount
    - Match Found (Yes/No)
    """
    data = []
    for match in matches:
        data.append({
            "Selection ID": match['Selection ID'],
            "Excel Amount": match['Selection Amount'],
            "PDF Filename": match['PDF Name'],
            "PDF Amount": match['PDF Amount'],
            "Match Found": "Yes" if match['Match Type'] == 'Exact' else "No"
        })
    
    df = pd.DataFrame(data)
    df.to_excel(output_file, index=False)
    logger.info(f"Results saved to {output_file}")

class PDFHandler:
    """Centralized PDF handling class with improved text extraction"""
    def __init__(self):
        self.output_dir = ensure_output_directory()
        self.temp_dir = self.output_dir / "temp"

    def extract_text(self, pdf_file: Union[str, Path, st.runtime.uploaded_file_manager.UploadedFile]) -> str:
        """Extract text from PDF file"""
        try:
            file_path = CommonUtils.get_file_path(pdf_file)
            with pdfplumber.open(file_path) as pdf:
                text = ""
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                return text
        except Exception as e:
            logger.error(f"Error extracting text from PDF: {str(e)}")
            return ""
        finally:
            if isinstance(pdf_file, st.runtime.uploaded_file_manager.UploadedFile):
                file_path.unlink(missing_ok=True)

    def extract_text_with_metadata(self, pdf_file: Union[str, Path, st.runtime.uploaded_file_manager.UploadedFile]) -> Dict[str, Any]:
        """Extract text and metadata from PDF file"""
        try:
            file_path = CommonUtils.get_file_path(pdf_file)
            with pdfplumber.open(file_path) as pdf:
                metadata = {
                    'text': '',
                    'invoice_number': None,
                    'date': None,
                    'project_number': None,
                    'po_number': None,
                    'total_amount': None,
                    'customer': None
                }
                
                text = ""
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                        
                        # Look for specific patterns
                        if 'Invoice Number:' in page_text:
                            metadata['invoice_number'] = self._extract_value(page_text, 'Invoice Number:', r'\d+')
                        if 'Project Number:' in page_text:
                            metadata['project_number'] = self._extract_value(page_text, 'Project Number:', r'\d+-[A-Z0-9]+')
                        if 'PO Number:' in page_text:
                            metadata['po_number'] = self._extract_value(page_text, 'PO Number:', r'P\d+-\d+')
                        if 'Total Due:' in page_text:
                            metadata['total_amount'] = self._extract_amount(page_text, 'Total Due:')
                
                metadata['text'] = text
                return metadata
        except Exception as e:
            logger.error(f"Error extracting text from PDF: {str(e)}")
            return {'text': '', 'error': str(e)}

    @staticmethod
    def _extract_value(text: str, prefix: str, pattern: str) -> Optional[str]:
        """Extract value using regex pattern"""
        import re
        try:
            line = [l for l in text.split('\n') if prefix in l][0]
            match = re.search(pattern, line)
            return match.group(0) if match else None
        except:
            return None

    @staticmethod
    def _extract_amount(text: str, prefix: str) -> Optional[float]:
        """Extract currency amount"""
        import re
        try:
            line = [l for l in text.split('\n') if prefix in l][0]
            match = re.search(r'\$[\d,]+\.\d{2}', line)
            if match:
                return float(match.group(0).replace('$', '').replace(',', ''))
            return None
        except:
            return None

    def convert_to_images(self, pdf_file: Union[str, Path, st.runtime.uploaded_file_manager.UploadedFile], 
                         selection_id: str) -> List[str]:
        """Convert PDF pages to images"""
        try:
            output_dir = ensure_output_directory()
            file_path = self.get_file_path(pdf_file)
            
            # Create safe selection ID for filenames
            safe_id = str(selection_id).replace('.', '_').replace(' ', '_')
            
            # Convert PDF to images
            images = convert_from_path(str(file_path), poppler_path=r'C:\Program Files\poppler\poppler-24.08.0\Library\bin')
            
            image_paths = []
            for i, image in enumerate(images):
                image_path = output_dir / f"{safe_id}_page_{i+1}.png"
                image.save(str(image_path.resolve()), "PNG")
                logger.info(f"Saved PDF page {i+1} to: {image_path}")
                
                # Create resized version for Excel
                resized_path = resize_pdf_image(image_path)
                image_paths.append(str(resized_path))
            
            return image_paths
        except Exception as e:
            logger.error(f"Error converting PDF to images: {str(e)}")
            return []
        finally:
            if isinstance(pdf_file, st.runtime.uploaded_file_manager.UploadedFile):
                try:
                    file_path.unlink(missing_ok=True)
                except Exception as e:
                    logger.error(f"Error cleaning up temp file: {e}")

    def cleanup(self) -> None:
        """Clean up temporary files and images"""
        try:
            if self.temp_dir.exists():
                shutil.rmtree(self.temp_dir)
            self.temp_dir.mkdir(parents=True, exist_ok=True)
            
            # Only remove PNG files from output_dir
            for f in self.output_dir.glob("*.png"):
                f.unlink()
        except Exception as e:
            logger.error(f"Error during cleanup: {e}")

    def get_file_path(self, pdf_file: Union[str, Path, st.runtime.uploaded_file_manager.UploadedFile]) -> Path:
        """Get filesystem path for PDF file"""
        if isinstance(pdf_file, st.runtime.uploaded_file_manager.UploadedFile):
            # Save uploaded file to temp directory
            self.temp_dir.mkdir(parents=True, exist_ok=True)
            temp_path = self.temp_dir / pdf_file.name
            with open(temp_path, 'wb') as f:
                f.write(pdf_file.getvalue())
            return temp_path
        return Path(pdf_file)

class CommonUtils:
    """Shared utility functions used across modules"""
    @staticmethod
    def format_currency(amount: Union[float, str, Decimal]) -> str:
        try:
            if isinstance(amount, str):
                amount = float(amount.replace(',', ''))
            return f"${Decimal(str(amount)):,.2f}"
        except (InvalidOperation, ValueError) as e:
            log_error(f"Error formatting currency: {str(e)}")
            return "Invalid Amount"

    @staticmethod
    def normalize_amount(amount: Union[str, float, Decimal]) -> Decimal:
        """
        Normalize financial amounts into a consistent decimal format.
        
        Args:
            amount: The amount to normalize, can be string ("$1,234.56"), float (1234.56), or Decimal
            
        Returns:
            Decimal: Normalized amount with 2 decimal places
            
        Examples:
            >>> normalize_amount("$1,234.56")
            Decimal('1234.56')
            >>> normalize_amount(1234.56)
            Decimal('1234.56')
            >>> normalize_amount("1,234")
            Decimal('1234.00')
            >>> normalize_amount("invalid")
            Decimal('0.00')
        """
        try:
            # Handle None or empty input
            if amount is None or (isinstance(amount, str) and not amount.strip()):
                logger.debug("Empty or None amount received")
                return Decimal('0.00')

            # Convert to string if float or Decimal
            if isinstance(amount, (float, Decimal)):
                amount_str = str(amount)
            else:
                amount_str = str(amount).strip()

            # Remove currency symbols, commas, and spaces
            cleaned = amount_str.replace('$', '').replace(',', '').replace(' ', '')

            # Convert to Decimal and quantize to 2 decimal places
            result = Decimal(cleaned).quantize(Decimal('0.01'))

            # Validate result is not negative
            if result < 0:
                logger.warning(f"Negative amount normalized: {result}")

            return result

        except (InvalidOperation, ValueError, TypeError) as e:
            logger.error(f"Error normalizing amount '{amount}': {str(e)}")
            return Decimal('0.00')
        except Exception as e:
            logger.error(f"Unexpected error normalizing amount '{amount}': {str(e)}")
            return Decimal('0.00')

    @staticmethod
    def get_file_path(file: Union[str, Path, st.runtime.uploaded_file_manager.UploadedFile]) -> Path:
        if isinstance(file, st.runtime.uploaded_file_manager.UploadedFile):
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp:
                tmp.write(file.getvalue())
                return Path(tmp.name)
        return Path(file)

    @staticmethod
    def log_error(message: str, exception: Optional[Exception] = None) -> None:
        """Log error message with optional exception details"""
        logger = logging.getLogger(__name__)
        logger.error(message)
        if exception:
            logger.error(f"Exception details: {str(exception)}", exc_info=True)

def create_selection_folders(selections: List[str]) -> Dict[str, Path]:
    """Create folders for each selection"""
    base_path = Path(OUTPUT_FOLDER)
    base_path.mkdir(exist_ok=True)
    
    folders = {}
    for selection_id in selections:
        folder = base_path / str(selection_id)
        folder.mkdir(exist_ok=True)
        folders[selection_id] = folder
    
    return folders

def create_support_summary(matches: List[Dict], output_path: Path) -> None:
    """Create Excel summary with support categories"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Support Summary"
    
    # Write headers
    for col, header in enumerate(EXCEL_HEADERS, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Write data rows
    for row, match in enumerate(matches, 2):
        selection_data = match['Selection Data']
        ws.cell(row=row, column=1, value=match['Selection ID'])
        ws.cell(row=row, column=2, value=selection_data.get('Amount', ''))
        ws.cell(row=row, column=3, value=selection_data.get('Description', ''))
        
        # Mark support categories
        pdf_type = match.get('PDF Type', '')
        if pdf_type:
            col = EXCEL_HEADERS.index(pdf_type) + 1
            ws.cell(row=row, column=col, value='x')
    
    wb.save(output_path)

def create_preview_table(matches: List[Dict], output_path: Path) -> pd.DataFrame:
    """Create preview table showing support status"""
    data = []
    for match in matches:
        row = {
            'Selection ID': match['Selection ID'],
            'Amount': match['Selection Amount']
        }
        # Add column for each category
        for category in SUPPORT_CATEGORIES:
            row[category] = 'x' if match.get('PDF Type') == category else ''
        data.append(row)
    
    df = pd.DataFrame(data)
    df.to_excel(output_path, index=False)
    return df